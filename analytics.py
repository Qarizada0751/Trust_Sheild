import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import plotly.express as px
from sqlalchemy import create_engine, text
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill, Alignment

pd.options.mode.chained_assignment = None

load_dotenv()
DB_USER = os.getenv("DB_USER", "mi_user")
DB_PASS = os.getenv("DB_PASS", "strongpass")
DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = os.getenv("DB_PORT", "5432")
DB_NAME = os.getenv("DB_NAME", "mercadoinsights_db")
CONN_STR = f"postgresql+psycopg2://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}"
engine = create_engine(CONN_STR, future=True)

ROOT = os.getcwd()
CHARTS_DIR = os.path.join(ROOT, "charts")
EXPORTS_DIR = os.path.join(ROOT, "exports")
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

plt.style.use("seaborn-v0_8")

def load_queries(path="queries.sql"):
    queries = {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = f.read()
    except FileNotFoundError:
        return queries
    blocks = raw.split("-- Q_")
    for b in blocks:
        b = b.strip()
        if not b: continue
        lines = b.splitlines()
        header = lines[0].strip()
        key = header.split()[0].split(":")[0].strip()
        sql = "\n".join(lines[1:]).strip()
        if key and sql:
            queries[key] = sql
    return queries

queries = load_queries("queries.sql")

def run_query(sql):
    try:
        return pd.read_sql_query(text(sql), engine)
    except Exception as e:
        print(f"[WARN] SQL failed: {e}")
        return pd.DataFrame()

def save_msg(path): print(f"  → saved: {path}")


def chart_pie():
    df = run_query(queries.get("PIE", "SELECT NULL LIMIT 0"))
    if df.empty: return pd.DataFrame()
    if "orders_count" in df.columns and "payments_count" not in df.columns:
        df = df.rename(columns={"orders_count": "payments_count"})

    total = df["payments_count"].sum()
    df["pct"] = df["payments_count"] / total * 100
    small = df[df["pct"] < 10]
    big = df[df["pct"] >= 10]
    if not small.empty:
        other_sum = small["payments_count"].sum()
        big = pd.concat([big, pd.DataFrame([{"payment_type": "Other", "payments_count": other_sum}])], ignore_index=True)
    df_plot = big.sort_values("payments_count", ascending=False)

    fig, ax = plt.subplots(figsize=(8, 8))
    wedges, texts, autotexts = ax.pie(df_plot["payments_count"], labels=df_plot["payment_type"],
                                      autopct="%1.1f%%", startangle=140, textprops={"weight": "bold"})
    ax.set_title("Orders by Payment Type (share)", pad=14)
    legend_labels = [f"{r.payment_type}: {int(r.payments_count):,} orders" for _, r in df_plot.iterrows()]
    ax.legend(wedges, legend_labels, title="Counts", loc="center left", bbox_to_anchor=(1, 0.5))
    path = os.path.join(CHARTS_DIR, "01_pie_payment_type.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ Pie chart saved →", path)
    return df

def chart_bar():
    df = run_query(queries.get("BAR", "SELECT NULL LIMIT 0"))
    if df.empty: return pd.DataFrame()
    df = df.sort_values("revenue", ascending=False)
    fig, ax = plt.subplots(figsize=(12, 6))
    categories = df["product_category_name"].str.replace("_", " ").str.title()
    bars = ax.bar(categories, df["revenue"], color="#2b8cbe", edgecolor="black")
    ax.set_title("Top Product Categories by Revenue", pad=14)
    ax.set_ylabel("Revenue (BRL)"); ax.set_xlabel("Category"); ax.tick_params(axis="x", rotation=40)
    for b, val in zip(bars, df["revenue"]):
        ax.text(b.get_x() + b.get_width()/2, val, f"{val:,.0f}", ha="center", va="bottom", fontsize=9)
    path = os.path.join(CHARTS_DIR, "02_bar_top_categories.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ Bar chart saved →", path)
    return df

def chart_hbar():
    df = run_query(queries.get("HBAR", "SELECT NULL LIMIT 0"))
    if df.empty: return pd.DataFrame()
    df = df.sort_values("avg_delivery_days")
    fig, ax = plt.subplots(figsize=(9, 6))
    bars = ax.barh(df["customer_state"], df["avg_delivery_days"], color="#66c2a5", edgecolor="black")
    ax.set_xlabel("Average delivery days"); ax.set_title("Average Delivery Time by State (days)", pad=14)
    for i, v in enumerate(df["avg_delivery_days"]):
        ax.text(v + 0.2, i, f"{v:.1f}", va="center", fontsize=9)
    path = os.path.join(CHARTS_DIR, "03_hbar_delivery_by_state.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ HBar chart saved →", path)
    return df

def chart_line():
    df = run_query(queries.get("LINE", "SELECT NULL LIMIT 0"))
    if df.empty: return pd.DataFrame()
    df["month"] = pd.to_datetime(df["month"])
    df = df.sort_values("month")
    df["ma3"] = df["revenue"].rolling(3, min_periods=1).mean()
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.plot(df["month"], df["revenue"], marker="o", label="Monthly revenue")
    ax.plot(df["month"], df["ma3"], linestyle="--", label="3-mo MA")
    ax.set_title("Monthly Revenue Trend", pad=14); ax.set_xlabel("Month"); ax.set_ylabel("Revenue (BRL)")
    ax.legend(); plt.xticks(rotation=40)
    path = os.path.join(CHARTS_DIR, "04_line_monthly_revenue.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ Line chart saved →", path)
    return df

def chart_hist():
    df = run_query(queries.get("HIST", "SELECT 1 as dummy"))
    if df.empty: return pd.DataFrame()
    fig, ax = plt.subplots(figsize=(10, 6))

    if "price" in df.columns:
        ax.hist(df["price"], bins=30, color="skyblue", edgecolor="black")
        mean = df["price"].mean()
        ax.axvline(mean, color="red", linestyle="--", label=f"Mean: {mean:.2f}")
        ax.set_title("Product Price Distribution", fontsize=16, pad=20)
        ax.set_xlabel("Price (BRL)"); ax.set_ylabel("Frequency"); ax.legend()
    elif "price_segment" in df.columns and "items_count" in df.columns:
        df = df.sort_values("items_count", ascending=False)
        bars = ax.bar(df["price_segment"], df["items_count"], color="skyblue", edgecolor="black")
        for b, val in zip(bars, df["items_count"]):
            ax.text(b.get_x() + b.get_width()/2, val, f"{val:,}", ha="center", va="bottom", fontsize=9)
        ax.set_title("Product Price Segments", fontsize=16, pad=20)
        ax.set_xlabel("Price Segment"); ax.set_ylabel("Number of Items")
        plt.xticks(rotation=30)
    else:
        return pd.DataFrame()

    path = os.path.join(CHARTS_DIR, "05_hist_price_segments.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ Histogram/Segments chart saved →", path)
    return df

def chart_scatter():
    df = run_query(queries.get("SCATTER", "SELECT NULL LIMIT 0"))
    if df.empty: return pd.DataFrame()
    sample = df.dropna(subset=["price", "freight_value"]).sample(n=min(4000, len(df)), random_state=1)
    fig, ax = plt.subplots(figsize=(10, 6))
    sc = ax.scatter(sample["price"], sample["freight_value"], c=sample["price"], cmap="viridis", alpha=0.6, s=18)
    corr = sample["price"].corr(sample["freight_value"])
    ax.set_title(f"Price vs Freight (corr={corr:.2f})", pad=14)
    ax.set_xlabel("Price (BRL)"); ax.set_ylabel("Freight (BRL)")
    plt.colorbar(sc, ax=ax).set_label("Price (BRL)")
    path = os.path.join(CHARTS_DIR, "06_scatter_price_freight.png")
    fig.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print("✔ Scatter chart saved →", path)
    return df


def export_to_excel_advanced(sheets: dict, filename="mercadoinsights_report.xlsx"):
    filepath = os.path.join(EXPORTS_DIR, filename)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df.empty: continue
            safe_name = name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False, startrow=0, startcol=0)

    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        ws.freeze_panes = "B2"

 
        header_row = 1
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=header_row, column=col)
            c.fill = header_fill
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        last_col, last_row = ws.max_column, ws.max_row
        if last_row >= 2 and last_col >= 1:
            rng = f"A2:{get_column_letter(last_col)}{last_row}"
            rule = ColorScaleRule(start_type="min", start_color="63BE7B",
                                  mid_type="percentile", mid_value=50, mid_color="FFDD71",
                                  end_type="max", end_color="F8696B")
            ws.conditional_formatting.add(rng, rule)

        ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
    print(f"✔ Excel report created: {filepath}")
    return filepath

def main():
    print("Running analytics pipeline — creating charts + Excel export")
    results = {}
    results["Payment_Summary"] = chart_pie()
    results["Top_Categories"] = chart_bar()
    results["Delivery_by_State"] = chart_hbar()
    results["Monthly_Revenue"] = chart_line()
    results["Price_Distribution"] = chart_hist()
    results["Scatter_Raw"] = chart_scatter()
    export_to_excel_advanced(results, "mercadoinsights_report.xlsx")

if __name__ == "__main__":
    main()
